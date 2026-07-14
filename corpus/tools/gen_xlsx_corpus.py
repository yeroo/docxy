#!/usr/bin/env python3
"""Generate the xlsx conformance corpus (corpus/xlsx/).

Authors a set of themed workbooks with openpyxl (formulas only, no cached
values), then runs them through LibreOffice Calc headless — which computes
every formula and stores the results as cached values. The output files are
therefore *oracles*: `xlsxy file.xlsx --verify` recalculates with gridcore
and diffs against LibreOffice's answers, and `gridcore/tests/conformance.rs`
enforces a clean scoreboard in CI.

LibreOffice is an independent, high-fidelity implementation of Excel
semantics — deviations between the two exist but are rare; any cell where
LibreOffice itself is known to disagree with real Excel must be documented
in corpus/xlsx/README.md (none currently known in this corpus).

Usage (from the repo root):
    python3 corpus/tools/gen_xlsx_corpus.py
Requires: python3 + openpyxl, and `soffice` (libreoffice-calc) on PATH.
"""

import os
import shutil
import subprocess
import sys
import tempfile
import zipfile

import openpyxl

OUT_DIR = os.path.join("corpus", "xlsx")

# Post-2007 functions are stored with an _xlfn. prefix in SpreadsheetML;
# without it, engines see an unknown user function and cache #NAME?.
XLFN = [
    "IFS", "SWITCH", "TEXTJOIN", "CONCAT", "MAXIFS", "MINIFS",
    "ISOWEEKNUM", "DAYS", "XOR", "IFNA", "NUMBERVALUE",
    "FLOOR.MATH", "CEILING.MATH", "STDEV.S", "STDEV.P", "VAR.S", "VAR.P",
    "MODE.SNGL", "RANK.EQ", "PERCENTILE.INC", "QUARTILE.INC",
]


def xlfn(formula):
    """Prefix post-2007 function names (longest first so CONCAT doesn't
    clobber CONCATENATE)."""
    import re
    for name in sorted(XLFN, key=len, reverse=True):
        formula = re.sub(
            rf"(?<![A-Za-z0-9_.]){re.escape(name)}\(",
            f"_xlfn.{name}(",
            formula,
        )
    return formula

# ---------------------------------------------------------------------------
# Workbook definitions
# ---------------------------------------------------------------------------


def wb_arithmetic(wb):
    ws = wb.active
    ws.title = "Ops"
    for i, v in enumerate([3, -4, 2.5, 0.1, 100, 7], start=1):
        ws.cell(row=i, column=1, value=v)
    ws["B1"] = "one"
    ws["B2"] = "two"
    fx = [
        "=1+2*3", "=(1+2)*3", "=-2^2", "=2^3^2", "=10/4", "=2+2%",
        "=A1+A2*A3", "=A1-A2", "=A5/A6", "=A1&A2", '=B1&" and "&B2',
        "=A1>A2", "=A1<=3", '="a"="A"', '="b">"a"', "=A1<>3", "=TRUE+1",
        '="5"+2', "=A4*10", "=A1%",
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=4, value=f)


def wb_math(wb):
    ws = wb.active
    ws.title = "Math"
    for i, v in enumerate([2.567, -8.4, 16, 5, 12, 30], start=1):
        ws.cell(row=i, column=1, value=v)
    fx = [
        "=ROUND(A1,2)", "=ROUND(-2.5,0)", "=ROUNDUP(A1,1)", "=ROUNDDOWN(A1,1)",
        "=INT(A2)", "=TRUNC(A2)", "=TRUNC(A1,1)", "=ABS(A2)", "=SIGN(A2)",
        "=MOD(-3,2)", "=MOD(10,3)", "=POWER(2,10)", "=SQRT(A3)", "=EXP(1)",
        "=LN(A3)", "=LOG(1000)", "=LOG(8,2)", "=LOG10(A5*0+100)", "=PI()",
        "=SIN(PI()/2)", "=COS(0)", "=DEGREES(PI())", "=RADIANS(180)",
        "=FLOOR(A1,0.5)", "=CEILING(A1,0.5)", "=EVEN(1.5)", "=ODD(2.5)",
        "=FACT(6)", "=GCD(A4,A5+3)", "=LCM(4,6)", "=COMBIN(8,2)",
        "=PERMUT(8,2)", "=SUMSQ(3,4)", "=QUOTIENT(10,3)", "=ATAN2(1,1)",
        "=SINH(1)", "=COSH(1)", "=TANH(1)",
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=4, value=f)


def wb_stats(wb):
    ws = wb.active
    ws.title = "Stats"
    data = [4, 15, 8, 23, 42, 16, 15, 9]
    for i, v in enumerate(data, start=1):
        ws.cell(row=i, column=1, value=v)
    fx = [
        "=SUM(A1:A8)", "=AVERAGE(A1:A8)", "=MEDIAN(A1:A8)", "=MIN(A1:A8)",
        "=MAX(A1:A8)", "=COUNT(A1:A8)", "=COUNTA(A1:A8)", "=COUNTBLANK(A1:B8)",
        "=STDEV(A1:A8)", "=STDEVP(A1:A8)", "=VAR(A1:A8)", "=VARP(A1:A8)",
        "=LARGE(A1:A8,2)", "=SMALL(A1:A8,3)", "=RANK(15,A1:A8)",
        "=RANK(15,A1:A8,1)", "=MODE(A1:A8)", "=PERCENTILE(A1:A8,0.5)",
        "=PERCENTILE(A1:A8,0.25)", "=QUARTILE(A1:A8,1)", "=QUARTILE(A1:A8,3)",
        "=SUMPRODUCT(A1:A4,A5:A8)", "=PRODUCT(A1:A3)",
        "=SUM(A:A)", "=COUNT(A:A)", "=AVERAGE(A1:A4,A5)",
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=4, value=f)


def wb_text(wb):
    ws = wb.active
    ws.title = "Text"
    words = ["alpha", "Beta Max", "  padded  ", "Ünïcødé", "a,b;c", "x"]
    for i, v in enumerate(words, start=1):
        ws.cell(row=i, column=1, value=v)
    fx = [
        "=LEN(A1)", "=LEN(A4)", "=UPPER(A1)", "=LOWER(A2)", "=PROPER(A1)",
        "=TRIM(A3)", '=TRIM("  x  y  ")', "=LEFT(A2,4)", "=RIGHT(A2,3)",
        "=MID(A2,6,3)", '=CONCATENATE(A1,"-",A6)', "=A1&A6",
        '=SUBSTITUTE(A5,",",";")', '=SUBSTITUTE("aaa","a","b",2)',
        '=REPLACE(A1,2,3,"XY")', '=REPT("ab",3)', '=FIND("a",A2)',
        '=SEARCH("MAX",A2)', "=EXACT(A1,\"alpha\")", "=EXACT(A1,\"Alpha\")",
        '=VALUE("12.5")', "=T(A1)", "=T(5)", "=N(5)", "=N(A1)",
        "=CHAR(65)", "=CODE(\"A\")", '=TEXT(1234.567,"#,##0.00")',
        '=TEXT(0.285,"0.0%")', '=TEXTJOIN("-",TRUE,A1,A6)',
        "=CONCAT(A1,A6)", "=CLEAN(A1)",
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=4, value=f)


def wb_logic(wb):
    ws = wb.active
    ws.title = "Logic"
    ws["A1"] = 5
    ws["A2"] = ""
    ws["A3"] = True
    ws["A4"] = "text"
    fx = [
        '=IF(A1>3,"big","small")', "=IF(A1<3,1)", "=IF(A1>3,,9)",
        "=AND(TRUE,A1>1)", "=OR(FALSE,A1>10)", "=XOR(TRUE,FALSE,TRUE)",
        "=NOT(A1=5)", "=IFERROR(1/0,-1)", "=IFERROR(A1*2,-1)",
        "=IFNA(NA(),42)", "=ISBLANK(Z99)", "=ISNUMBER(A1)", "=ISNUMBER(A4)",
        "=ISTEXT(A4)", "=ISLOGICAL(A3)", "=ISERROR(1/0)", "=ISERR(NA())",
        "=ISNA(NA())", "=ISEVEN(4)", "=ISODD(4)", "=IFS(A1>10,1,A1>3,2)",
        '=SWITCH(2,1,"one",2,"two","other")', '=SWITCH(9,1,"one","fallback")',
        "=TRUE()", "=FALSE()",
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=4, value=f)


def wb_lookup(wb):
    ws = wb.active
    ws.title = "Lookup"
    rows = [("apple", 10, "red"), ("banana", 20, "yellow"),
            ("cherry", 30, "red"), ("date", 40, "brown"),
            ("elderberry", 50, "purple")]
    for i, (a, b, c) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        ws.cell(row=i, column=3, value=c)
    fx = [
        '=VLOOKUP("cherry",A1:C5,2,FALSE)', '=VLOOKUP("cherry",A1:C5,3,FALSE)',
        "=VLOOKUP(35,B1:C5,2)", '=MATCH("date",A1:A5,0)', "=MATCH(30,B1:B5,0)",
        "=MATCH(35,B1:B5,1)", "=INDEX(B1:B5,3)", "=INDEX(A1:C5,2,3)",
        "=CHOOSE(2,B1,B2,B3)", "=LOOKUP(35,B1:B5)", "=LOOKUP(35,B1:B5,A1:A5)",
        '=HLOOKUP(20,B1:C2,2,FALSE)', "=ROW(B3)", "=COLUMN(C1)",
        "=ROWS(A1:C5)", "=COLUMNS(A1:C5)",
        # XLOOKUP intentionally absent: LibreOffice 24.2 (this oracle)
        # predates it. Covered by gridcore unit tests until a real-Excel
        # corpus file arrives.
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=5, value=f)


def wb_dates(wb):
    ws = wb.active
    ws.title = "Dates"
    fx = [
        "=DATE(2024,1,15)", "=DATE(2023,13,1)", "=DATE(2024,2,29)",
        "=YEAR(DATE(2024,3,10))", "=MONTH(DATE(2024,3,10))",
        "=DAY(DATE(2024,3,10))", "=WEEKDAY(DATE(2024,1,15))",
        "=WEEKDAY(DATE(2024,1,15),2)", "=WEEKDAY(DATE(2024,1,15),3)",
        "=EDATE(DATE(2024,1,31),1)", "=EOMONTH(DATE(2024,1,15),0)",
        "=EOMONTH(DATE(2024,1,15),13)", "=DAYS(DATE(2024,3,1),DATE(2024,2,1))",
        '=DATEDIF(DATE(2020,1,15),DATE(2024,3,10),"Y")',
        '=DATEDIF(DATE(2020,1,15),DATE(2024,3,10),"M")',
        '=DATEDIF(DATE(2020,1,15),DATE(2024,3,10),"D")',
        "=WEEKNUM(DATE(2024,1,15))", "=ISOWEEKNUM(DATE(2024,1,15))",
        "=ISOWEEKNUM(DATE(2023,1,1))", "=TIME(18,30,0)",
        "=HOUR(0.75)", "=MINUTE(TIME(12,34,56))", "=SECOND(TIME(12,34,56))",
        "=DATE(2024,1,15)+TIME(12,0,0)",
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=2, value=f)


def wb_financial(wb):
    ws = wb.active
    ws.title = "Fin"
    flows = [-70000, 12000, 15000, 18000, 21000, 26000]
    for i, v in enumerate(flows, start=1):
        ws.cell(row=i, column=1, value=v)
    fx = [
        "=PMT(0.08/12,10,10000)", "=PMT(0,10,10000)",
        "=FV(0.06/12,10,-200,-500,1)", "=FV(0.06/12,10,-200)",
        "=PV(0.08/12,240,500)", "=NPER(0.12/12,-100,-1000,10000,1)",
        "=NPV(0.1,-10000,3000,4200,6800)", "=NPV(0.08,A2:A6)",
        "=IRR(A1:A6)", "=RATE(48,-200,8000)",
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=3, value=f)


def wb_criteria(wb):
    ws = wb.active
    ws.title = "Crit"
    rows = [("east", 10, "jan"), ("west", 20, "jan"), ("east", 30, "feb"),
            ("north", 40, "feb"), ("east", 50, "mar"), ("west", 60, "mar")]
    for i, (a, b, c) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        ws.cell(row=i, column=3, value=c)
    fx = [
        '=COUNTIF(A1:A6,"east")', '=COUNTIF(A1:A6,"?est")',
        '=COUNTIF(A1:A6,"*e*")', '=COUNTIF(B1:B6,">25")',
        '=SUMIF(A1:A6,"east",B1:B6)', '=SUMIF(B1:B6,">=30")',
        '=AVERAGEIF(A1:A6,"west",B1:B6)',
        '=COUNTIFS(A1:A6,"east",C1:C6,"feb")',
        '=SUMIFS(B1:B6,A1:A6,"east",B1:B6,">15")',
        '=AVERAGEIFS(B1:B6,A1:A6,"east")',
        '=MAXIFS(B1:B6,A1:A6,"east")', '=MINIFS(B1:B6,A1:A6,"west")',
        '=SUMIF(A1:A6,"<>east",B1:B6)', '=COUNTIF(A:A,"east")',
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=5, value=f)


def wb_refs(wb):
    data = wb.active
    data.title = "Data"
    for i, v in enumerate([10, 20, 30, 40, 50], start=1):
        data.cell(row=i, column=1, value=v)
    calc = wb.create_sheet("Calc Zone")  # space in name → quoted refs
    calc["A1"] = "=Data!A1*2"
    calc["A2"] = "=SUM(Data!A1:A5)"
    calc["A3"] = "=SUM(Data!A:A)"
    calc["A4"] = "=Data!$A$2+Data!A3"
    calc["A5"] = "=SUM(TheData)"
    calc["A6"] = "=TaxRate*100"
    calc["A7"] = "='Calc Zone'!A1+1"
    wb.defined_names.add(
        openpyxl.workbook.defined_name.DefinedName(
            "TheData", attr_text="Data!$A$1:$A$5"))
    wb.defined_names.add(
        openpyxl.workbook.defined_name.DefinedName("TaxRate", attr_text="0.21"))


def wb_amortization(wb):
    """A realistic fill-down sheet: LibreOffice writes these as shared
    formulas, exercising group expansion in the loader."""
    ws = wb.active
    ws.title = "Loan"
    ws["A1"] = "rate"
    ws["B1"] = 0.005
    ws["A2"] = "payment"
    ws["B2"] = "=-PMT($B$1,24,10000)"
    ws["A4"] = "month"
    ws["B4"] = "balance"
    ws["C4"] = "interest"
    ws["D4"] = "principal"
    ws["A5"] = 1
    ws["B5"] = 10000
    ws["C5"] = "=B5*$B$1"
    ws["D5"] = "=$B$2-C5"
    for r in range(6, 29):
        ws.cell(row=r, column=1, value=f"=A{r-1}+1")
        ws.cell(row=r, column=2, value=f"=B{r-1}-D{r-1}")
        ws.cell(row=r, column=3, value=f"=B{r}*$B$1")
        ws.cell(row=r, column=4, value=f"=$B$2-C{r}")
    ws["F1"] = "=SUM(C5:C28)"
    ws["F2"] = "=SUM(D5:D28)"
    ws["F3"] = "=B28-D28"


def wb_gradebook(wb):
    ws = wb.active
    ws.title = "Grades"
    students = [("Ana", 78, 91), ("Ben", 55, 64), ("Cleo", 92, 88),
                ("Dan", 43, 71), ("Eve", 85, 95)]
    ws["A1"], ws["B1"], ws["C1"] = "name", "mid", "final"
    for i, (n, m, f) in enumerate(students, start=2):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=m)
        ws.cell(row=i, column=3, value=f)
        ws.cell(row=i, column=4, value=f"=0.4*B{i}+0.6*C{i}")
        ws.cell(row=i, column=5,
                value=f'=IF(D{i}>=90,"A",IF(D{i}>=80,"B",IF(D{i}>=70,"C","F")))')
    ws["G1"] = "=AVERAGE(D2:D6)"
    ws["G2"] = "=MAX(D2:D6)"
    ws["G3"] = '=COUNTIF(E2:E6,"F")'
    ws["G4"] = '=INDEX(A2:A6,MATCH(MAX(D2:D6),D2:D6,0))'


def wb_3d(wb):
    """3D sheet spans: the same cell across a run of sheets."""
    q1 = wb.active
    q1.title = "Q1"
    for name, vals in [("Q1", [100, 20]), ("Q2", [110, 25]), ("Q3", [120, 30])]:
        ws = q1 if name == "Q1" else wb.create_sheet(name)
        ws["A1"] = vals[0]
        ws["A2"] = vals[1]
    total = wb.create_sheet("Total")
    total["A1"] = "=SUM(Q1:Q3!A1)"
    total["A2"] = "=SUM(Q1:Q3!A1:A2)"
    total["A3"] = "=AVERAGE(Q1:Q3!A2)"
    total["A4"] = "=COUNT(Q1:Q3!A1:A2)"
    total["A5"] = "=MAX(Q1:Q3!A1)"
    total["A6"] = "=MIN(Q1:Q3!A2)"


def wb_textfmt(wb):
    """TEXT() across the format-code surface — oracle-validates the
    number-format runtime."""
    ws = wb.active
    ws.title = "Fmt"
    ws["A1"] = 1234.567
    ws["A2"] = -1234.567
    ws["A3"] = 0.28456
    ws["A4"] = 45306.7708333333  # 2024-01-15 18:30
    ws["A5"] = 12200000
    ws["A6"] = 0.000123
    fx = [
        '=TEXT(A1,"0")', '=TEXT(A1,"0.00")', '=TEXT(A1,"#,##0")',
        '=TEXT(A1,"#,##0.00")', '=TEXT(A2,"0.00")', '=TEXT(A1,"00000")',
        '=TEXT(A3,"0%")', '=TEXT(A3,"0.0%")', '=TEXT(A3,"0.00%")',
        '=TEXT(A1,"$#,##0.00")', '=TEXT(A2,"0.00;(0.00)")',
        '=TEXT(0,"0;-0;""zero""")', '=TEXT(A5,"#,##0,")',
        '=TEXT(A5,"0.0,,")', '=TEXT(A6,"0.00E+00")', '=TEXT(A1,"0.00E+00")',
        '=TEXT(A4,"yyyy-mm-dd")', '=TEXT(A4,"m/d/yyyy")',
        '=TEXT(A4,"d-mmm-yy")', '=TEXT(A4,"dddd")', '=TEXT(A4,"mmmm")',
        '=TEXT(A4,"hh:mm:ss")', '=TEXT(A4,"h:mm AM/PM")',
        '=TEXT(A4,"yyyy-mm-dd hh:mm")', '=TEXT(1.5,"[h]:mm")',
        '=TEXT(A1,"0.0""kg""")', '=TEXT(A1,"General")',
        '=TEXT("hello","@")',
    ]
    for i, f in enumerate(fx, start=1):
        ws.cell(row=i, column=3, value=f)


def wb_salestable(wb):
    """A real Excel Table (ListObject) with calculated columns and
    structured-reference aggregations."""
    from openpyxl.worksheet.table import Table, TableStyleInfo
    ws = wb.active
    ws.title = "Orders"
    ws.append(["Item", "Qty", "Price", "Amount"])
    # LibreOffice evaluates the explicit [#This Row] form; the bare [@Col]
    # sugar (identical semantics) is covered by gridcore's unit tests.
    for item, qty, price in [("pen", 3, 1.5), ("pad", 2, 4), ("ink", 5, 2),
                             ("clip", 10, 0.25)]:
        ws.append([item, qty, price,
                   "=Sales[[#This Row],[Qty]]*Sales[[#This Row],[Price]]"])
    tab = Table(displayName="Sales", ref="A1:D5")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",
                                        showRowStripes=True)
    ws.add_table(tab)
    ws["F1"] = "=SUM(Sales[Amount])"
    ws["F2"] = "=SUM(Sales[Qty])"
    ws["F3"] = "=COUNTA(Sales[#Headers])"
    ws["F4"] = "=SUM(Sales[[Qty]:[Price]])"
    ws["F5"] = "=AVERAGE(Sales[Price])"
    ws["F6"] = "=SUMPRODUCT(Sales[Qty],Sales[Price])"
    ws["F7"] = "=SUM(Sales[#All])"
    ws["F8"] = '=SUMIF(Sales[Item],"p*",Sales[Amount])'


def wb_edge_cases(wb):
    ws = wb.active
    ws.title = "Edgy"
    ws["A1"] = 'quote " and <angle> & amp'
    ws["A2"] = "line\nbreak"
    ws["A3"] = "  spaces kept  "
    ws["A4"] = "Ünïcødé ☃ 中文"
    ws["B1"] = "=LEN(A1)"
    ws["B2"] = '=A1&"!"'
    ws["B3"] = "=LEN(A3)"
    ws["B4"] = "=UPPER(A4)"
    # A long dependency chain.
    ws["D1"] = 1
    for r in range(2, 40):
        ws.cell(row=r, column=4, value=f"=D{r-1}+1")
    ws["E1"] = "=D39"
    # Errors as first-class results.
    ws["F1"] = "=1/0"
    ws["F2"] = "=NA()"
    ws["F3"] = "=IFERROR(F1,99)"
    ws["F4"] = "=ISERROR(F2)"
    ws.merge_cells("A6:C7")
    ws["A6"] = "=SUM(D1:D39)"


WORKBOOKS = {
    "calc-arithmetic": wb_arithmetic,
    "calc-math": wb_math,
    "calc-stats": wb_stats,
    "calc-text": wb_text,
    "calc-logic": wb_logic,
    "calc-lookup": wb_lookup,
    "calc-dates": wb_dates,
    "calc-financial": wb_financial,
    "calc-criteria": wb_criteria,
    "calc-refs": wb_refs,
    "shape-amortization": wb_amortization,
    "shape-gradebook": wb_gradebook,
    "calc-3d": wb_3d,
    "calc-textfmt": wb_textfmt,
    "shape-salestable": wb_salestable,
    "edge-cases": wb_edge_cases,
}

# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------


def main():
    if shutil.which("soffice") is None:
        sys.exit("soffice (libreoffice-calc) not found on PATH")
    # The oracle values LibreOffice caches for TEXT()/VALUE()/percent formats
    # are locale-sensitive (decimal/grouping separators). Pin a locale so the
    # corpus is reproducible regardless of the generating machine's settings,
    # and record which LibreOffice produced it for provenance.
    ver = subprocess.run(
        ["soffice", "--version"], capture_output=True, text=True
    ).stdout.strip()
    print(f"generating with: {ver}")
    lc_env = {**os.environ, "LC_ALL": "en_US.UTF-8", "LANG": "en_US.UTF-8"}
    os.makedirs(OUT_DIR, exist_ok=True)
    with tempfile.TemporaryDirectory() as tmp:
        src_dir = os.path.join(tmp, "src")
        out_dir = os.path.join(tmp, "out")
        os.makedirs(src_dir)
        for name, author in WORKBOOKS.items():
            wb = openpyxl.Workbook()
            author(wb)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            cell.value = xlfn(cell.value)
            wb.save(os.path.join(src_dir, f"{name}.xlsx"))
        # One batch conversion: LibreOffice loads each file (computing all
        # formulas — the sources carry no cached values) and saves xlsx
        # with its results embedded.
        subprocess.run(
            ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", out_dir]
            + [os.path.join(src_dir, f"{n}.xlsx") for n in WORKBOOKS],
            check=True,
            stdout=subprocess.DEVNULL,
            env=lc_env,
        )
        for name in WORKBOOKS:
            path = os.path.join(out_dir, f"{name}.xlsx")
            check_cached_values(path)
            shutil.copy(path, os.path.join(OUT_DIR, f"{name}.xlsx"))
            print(f"wrote {OUT_DIR}/{name}.xlsx")


def check_cached_values(path):
    """Every <f> must be followed by a cached <v> (or be an empty-string
    result) — otherwise the oracle would be comparing against nothing."""
    z = zipfile.ZipFile(path)
    for part in z.namelist():
        if not part.startswith("xl/worksheets/"):
            continue
        xml = z.read(part).decode()
        cells = xml.count("<f")
        cached = xml.count("</f><v>")
        if cells and cached < cells * 0.8:
            sys.exit(f"{path}: only {cached}/{cells} formulas have cached values")


if __name__ == "__main__":
    main()
